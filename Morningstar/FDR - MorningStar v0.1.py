import os
import time
import subprocess
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import openpyxl

subprocess.check_call(["pip", "install", "xlrd"])
import xlrd

def create_subfolder(directory, subfolder):
    subfolder_path = os.path.join(directory, subfolder)
    if not os.path.exists(subfolder_path):
        os.makedirs(subfolder_path)
    return subfolder_path

def download_morningstar_excel(ticker):
    print(f"Downloading financial data for {ticker}...")
    # Set up the Selenium WebDriver
    driver = webdriver.Chrome()  # You need to have chromedriver installed and in your PATH
    driver.get(f"https://www.morningstar.com/stocks/xnas/{ticker}/financials")

    try:
        # Expand the detail view
        expand_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "more-data.sal-mds-link"))
        )
        expand_button.click()
        print(f"Expanded detail view for {ticker}.")

        # Click on "Export Data" button - Income Statement
        export_button_IS = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "sal-financials-details__exportSection"))
        )
        export_button_IS.click()
        time.sleep(5)  # Wait for the download to complete
        # Click the download button
        print(f"Download of the Income Statement of {ticker} completed.")

        # Create subfolder on desktop
        desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        subfolder_path = create_subfolder(desktop_path, ticker)

        # Move downloaded file to subfolder
        downloaded_file_IS = os.path.join(os.path.expanduser('~'), 'Downloads', "Income Statement_Annual_As Originally Reported.xls")
        new_file_path_IS = os.path.join(subfolder_path, f"{ticker}_Income_Statement.xlsx")
        os.rename(downloaded_file_IS, new_file_path_IS)

        # Convert xls to xlsx
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        with xlrd.open_workbook(new_file_path_IS) as wb:
            for sheet in wb.sheets():
                for row in sheet.get_rows():
                    worksheet.append([cell.value for cell in row])
        workbook.save(new_file_path_IS)

        # Locate the Balance Sheet button
        BS_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//button[@id='balanceSheet']"))
        )
        BS_button.click()
        # Click on "Export Data" button - Balance Sheet
        # Scroll down just a tiny bit after expanding detail view
        driver.execute_script("window.scrollBy(0, 75);")  # Adjust the value as needed
        time.sleep(5)
        export_button_BS = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "sal-financials-details__exportSection"))
        )
        export_button_BS.click()
        time.sleep(5)  # Wait for the download to complete
        # Click the download button
        print(f"Download of the Balance Sheet for {ticker} completed.")

        # Move downloaded file to subfolder
        downloaded_file_BS = os.path.join(os.path.expanduser('~'), 'Downloads', "Balance Sheet_Annual_As Originally Reported.xls")
        new_file_path_BS = os.path.join(subfolder_path, f"{ticker}_Balance_Sheet.xlsx")
        os.rename(downloaded_file_BS, new_file_path_BS)

        # Convert xls to xlsx
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        with xlrd.open_workbook(new_file_path_BS) as wb:
            for sheet in wb.sheets():
                for row in sheet.get_rows():
                    worksheet.append([cell.value for cell in row])
        workbook.save(new_file_path_BS)

        # Locate the Cash Flow button
        CF_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//button[@id='cashFlow']"))
        )
        CF_button.click()
        # Click on "Export Data" button - Cash Flow
        time.sleep(5)
        export_button_CF = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "sal-financials-details__exportSection"))
        )
        export_button_CF.click()
        # Click the download button
        time.sleep(2)  # Wait for the download to complete
        print(f"Download of the Cash Flow for {ticker} completed.")

        # Move downloaded file to subfolder
        downloaded_file_CF = os.path.join(os.path.expanduser('~'), 'Downloads', "Cash Flow_Annual_As Originally Reported.xls")
        new_file_path_CF = os.path.join(subfolder_path, f"{ticker}_Cash_Flow.xlsx")
        os.rename(downloaded_file_CF, new_file_path_CF)

        # Convert xls to xlsx
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        with xlrd.open_workbook(new_file_path_CF) as wb:
            for sheet in wb.sheets():
                for row in sheet.get_rows():
                    worksheet.append([cell.value for cell in row])
        workbook.save(new_file_path_CF)

        return new_file_path_IS, new_file_path_BS, new_file_path_CF
    except Exception as e:
        print(f"Failed to download any financial data for {ticker}: {e}")
        return None, None, None
    finally:
        driver.quit()

def main():
    # List of tickers
    tickers = ['AAPL']  # Feel free to change the ticker of the company you want to analyse. However, do not add more than one tickers as the code will seemingly stop working normally.

    # Download financial data for each ticker
    for ticker in tickers:
        downloaded_files = []

        IS_path, BS_path, CF_path = download_morningstar_excel(ticker)
        if IS_path and BS_path and CF_path:
            downloaded_files.extend([IS_path, BS_path, CF_path])
            print(f"Download of financial data for {ticker} successful.")
        else:
            print(f"Download of financial data for {ticker} failed.")

        if downloaded_files:
            try:
                merged_file_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads', f"{ticker}_MS_financials.xlsx")
                with pd.ExcelWriter(merged_file_path) as writer:
                    for file_path in downloaded_files:
                        sheet_name = os.path.splitext(os.path.basename(file_path))[0]  # Extracting sheet name from file name
                        df = pd.read_excel(file_path)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"Downloaded documents for {ticker} merged successfully.")

                # Delete downloaded files
                for file_path in downloaded_files:
                    os.remove(file_path)
                print("Unused files deleted successfully.")
            except Exception as e:
                print(f"Failed to merge financial data for {ticker}: {e}")

if __name__ == "__main__":
    main()